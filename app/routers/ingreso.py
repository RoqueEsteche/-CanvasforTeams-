"""New-student/teacher onboarding: credential generation + Canvas/Teams creation + welcome email."""
import asyncio
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, File, UploadFile
from pydantic import BaseModel


def _err(exc: Exception) -> str:
    """Extract the most useful error message from any exception type."""
    return exc.detail if isinstance(exc, HTTPException) else str(exc)

from app.core.config import settings
from app.models.canvas import BulkResult
from app.services import canvas_client as canvas
from app.services import teams_client as graph
from app.services.credential_generator import generate_credentials
from app.services.email_service import send_welcome_email

router = APIRouter(prefix="/ingreso", tags=["Nuevo Ingreso"])
_ACCOUNT = settings.canvas_account_id


class StudentIn(BaseModel):
    full_name: str
    cedula: str
    personal_email: str
    role: str = "student"            # "student" | "teacher"
    platform: str = "both"           # "canvas" | "teams" | "both"
    program_type: str = "grado"      # "grado" | "mba" | "diplomado"
    program_name: str = ""           # nombre específico del diplomado/programa
    send_email: bool = True
    cc: list[str] = []               # extra CC recipients for this send


class BulkStudentsIn(BaseModel):
    students: list[StudentIn]


class ResendCredentialsIn(BaseModel):
    cedula: str
    personal_email: str
    full_name: str
    platform: str = "both"           # "canvas" | "teams" | "both"
    program_type: str = "grado"      # "grado" | "mba" | "diplomado"
    program_name: str = ""
    cc: list[str] = []


class BulkResendIn(BaseModel):
    students: list[ResendCredentialsIn]


class AccountCheckIn(BaseModel):
    cedula: str
    full_name: str = ""          # opcional, si se provee se genera el login_id para buscar
    platform: str = "both"       # "canvas" | "teams" | "both"


class BulkAccountCheckIn(BaseModel):
    students: list[AccountCheckIn]


class CredentialPreview(BaseModel):
    full_name: str
    cedula: str
    personal_email: str
    role: str
    login_id: str
    institutional_email: str
    password: str


def _resolve_login(creds: dict, role: str) -> tuple[str, str]:
    """Return (canvas_login_id, canvas_sis_user_id).

    Both students and teachers use the institutional email as login (unique_id)
    so they can authenticate with their email and password on both Canvas and Teams.
    The SIS user ID is always the cédula for institutional tracking.
    """
    return creds["email"], creds["cedula"]


@router.post("/preview", response_model=CredentialPreview, summary="Previsualizar credenciales")
async def preview_credentials(body: StudentIn):
    creds = generate_credentials(body.full_name, body.cedula, settings.institutional_domain)
    login_id, _ = _resolve_login(creds, body.role)
    return CredentialPreview(
        full_name=body.full_name,
        cedula=body.cedula,
        personal_email=body.personal_email,
        role=body.role,
        login_id=login_id,
        institutional_email=creds["email"],
        password=creds["password"],
    )


async def _canvas_user_exists(cedula: str, login_id: str) -> tuple[bool, dict]:
    """Check Canvas for existing user by SIS ID (cedula) or login_id.
    Returns (exists, user_info_dict).
    """
    # Primary: check by SIS user ID (cedula) — finds the user regardless of name changes
    try:
        user = await canvas.get(f"/users/sis_user_id:{cedula}")
        return True, {
            "found_by": "cedula",
            "canvas_id": user.get("id"),
            "name": user.get("name", ""),
            "login_id": user.get("login_id", ""),
            "email": user.get("email", ""),
        }
    except Exception:
        pass
    # Fallback: check by generated login_id (institutional email)
    try:
        search = await canvas.get(f"/accounts/{_ACCOUNT}/users", {"search_term": login_id})
        if isinstance(search, list):
            for u in search:
                if u.get("login_id", "").lower() == login_id.lower():
                    return True, {
                        "found_by": "login_id",
                        "canvas_id": u.get("id"),
                        "name": u.get("name", ""),
                        "login_id": u.get("login_id", ""),
                        "email": u.get("email", ""),
                    }
    except Exception:
        pass
    return False, {}


async def _teams_user_exists(upn: str) -> tuple[bool, dict]:
    """Check Azure AD for existing user by userPrincipalName.
    Returns (exists, user_info_dict).
    """
    try:
        user = await graph.get(f"/users/{upn}?$select=id,displayName,userPrincipalName,mail,accountEnabled,createdDateTime")
        return True, {
            "found_by": "upn",
            "azure_id": user.get("id"),
            "name": user.get("displayName", ""),
            "upn": user.get("userPrincipalName", ""),
            "mail": user.get("mail", ""),
            "account_enabled": user.get("accountEnabled"),
            "created": user.get("createdDateTime", ""),
        }
    except Exception as exc:
        err = str(exc)
        if "404" in err or "Request_ResourceNotFound" in err or "does not exist" in err.lower():
            return False, {}
        raise


async def _check_account(body: AccountCheckIn) -> dict[str, Any]:
    """Verify if a user exists in Canvas and/or Teams. Core logic used by single and bulk check."""
    result: dict[str, Any] = {
        "cedula": body.cedula,
        "full_name": body.full_name,
    }

    # Derive login_id from name if provided, else use cedula as search term
    if body.full_name.strip():
        creds = generate_credentials(body.full_name, body.cedula, settings.institutional_domain)
        login_id = creds["email"]
        upn = creds["email"]
        result["generated_email"] = creds["email"]
    else:
        login_id = body.cedula
        upn = ""

    if body.platform in ("canvas", "both"):
        try:
            exists, info = await _canvas_user_exists(body.cedula, login_id)
            result["canvas"] = {"exists": exists, **(info if exists else {})}
        except Exception as exc:
            result["canvas"] = {"exists": None, "error": _err(exc)}

    if body.platform in ("teams", "both"):
        if not upn:
            result["teams"] = {"exists": None, "error": "Se requiere nombre completo para buscar en Teams"}
        else:
            try:
                exists, info = await _teams_user_exists(upn)
                result["teams"] = {"exists": exists, **(info if exists else {})}
            except Exception as exc:
                result["teams"] = {"exists": None, "error": _err(exc)}

    return result


async def _create_student(student: StudentIn) -> dict[str, Any]:
    creds = generate_credentials(student.full_name, student.cedula, settings.institutional_domain)
    login_id, sis_user_id = _resolve_login(creds, student.role)
    results: dict[str, Any] = {
        "student": student.full_name,
        "role": student.role,
        "credentials": {**creds, "login_id": login_id},
    }

    if student.platform in ("canvas", "both"):
        try:
            # Pre-check: does user already exist by cedula or login?
            exists, info = await _canvas_user_exists(sis_user_id, login_id)
            if exists:
                results["canvas"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Canvas (registrado como: {info.get('name', '?')})",
                    "existing": info,
                }
            else:
                canvas_user = await canvas.post(
                    f"/accounts/{_ACCOUNT}/users",
                    {
                        "user": {"name": creds["full_name"], "short_name": creds["full_name"]},
                        "pseudonym": {
                            "unique_id": login_id,
                            "sis_user_id": sis_user_id,
                            "password": creds["password"],
                            "send_confirmation": False,
                        },
                        "communication_channel": {
                            "type": "email",
                            "address": creds["email"],
                            "skip_confirmation": True,
                        },
                    },
                )
                results["canvas"] = {"status": "ok", "id": canvas_user.get("id")}
        except Exception as exc:
            error_str = _err(exc)
            # Fallback reactive detection for any edge cases
            if "unique_id" in error_str.lower() or "taken" in error_str.lower() or "already" in error_str.lower():
                results["canvas"] = {"status": "exists", "error": f"Usuario ya existe en Canvas: {error_str}"}
            else:
                results["canvas"] = {"status": "error", "error": error_str}

    if student.platform in ("teams", "both"):
        parts = student.full_name.strip().split()
        sku = settings.azure_sku_teachers if student.role == "teacher" else settings.azure_sku_students
        try:
            # Pre-check: does user already exist in Azure AD by UPN?
            exists, info = await _teams_user_exists(creds["email"])
            if exists:
                results["teams"] = {
                    "status": "exists",
                    "error": f"Usuario ya existe en Teams (registrado como: {info.get('name', '?')})",
                    "existing": info,
                }
            else:
                az_user = await graph.post(
                    "/users",
                    {
                        "displayName": creds["full_name"],
                        "givenName": parts[0],
                        "surname": " ".join(parts[1:]) if len(parts) > 1 else "",
                        "userPrincipalName": creds["email"],
                        "mailNickname": creds["login_id"].replace(".", "_"),
                        "usageLocation": settings.usage_location,
                        "accountEnabled": True,
                        "passwordProfile": {
                            "forceChangePasswordNextSignIn": True,
                            "password": creds["password"],
                        },
                    },
                )
                await graph.assign_license(az_user["id"], sku)
                results["teams"] = {"status": "ok", "id": az_user.get("id"), "license": sku}
        except Exception as exc:
            error_str = _err(exc)
            # Fallback reactive detection
            if "ObjectConflict" in error_str or "already exists" in error_str.lower() or "conflictingObjects" in error_str:
                results["teams"] = {"status": "exists", "error": f"Usuario ya existe en Teams: {error_str}"}
            else:
                results["teams"] = {"status": "error", "error": error_str}

    if student.send_email:
        try:
            # Prepare attachments for diplomado
            attachments = []
            if student.program_type == "diplomado":
                template_dir = Path(__file__).parent.parent / "static" / "templates"
                pdf_files = [
                    template_dir / "2° Acceso a la Plataforma Teams- Instructivo.pdf",
                    template_dir / "3° Descargar grabacion en TEAMS - Instructivo.pdf",
                ]
                attachments = [str(f) for f in pdf_files if f.exists()]

            await send_welcome_email(
                to_email=student.personal_email,
                full_name=creds["full_name"],
                institutional_email=creds["email"],
                login_id=login_id,
                password=creds["password"],
                platform=student.platform,
                program_type=student.program_type,
                program_name=student.program_name,
                extra_cc=student.cc or None,
                attachments=attachments if attachments else None,
            )
            results["email"] = "sent"
        except Exception as exc:
            results["email"] = f"error: {exc}"

    return results


@router.post("/test-email", summary="Probar envío de correo")
async def test_email(
    to_email: str,
    program_type: str = "grado",
    program_name: str = "",
):
    """Envía un correo de prueba para verificar la configuración y la plantilla."""
    try:
        await send_welcome_email(
            to_email=to_email,
            full_name="Usuario de Prueba",
            institutional_email=f"prueba@{settings.institutional_domain}",
            login_id=f"prueba@{settings.institutional_domain}",
            password="Test-Pw123",
            platform="both",
            program_type=program_type,
            program_name=program_name,
        )
        return {"status": "ok", "to": to_email, "from": settings.smtp_from, "program_type": program_type}
    except Exception as exc:
        return {"status": "error", "detail": str(exc)}


@router.post("/create", summary="Crear credenciales para un alumno o docente")
async def create_student(body: StudentIn):
    return await _create_student(body)


@router.post("/bulk", summary="Crear credenciales masivas")
async def create_students_bulk(body: BulkStudentsIn) -> BulkResult:
    result = BulkResult()

    async def _run(student: StudentIn):
        try:
            data = await _create_student(student)
            result.succeeded.append(data)
        except Exception as exc:
            result.failed.append({"student": student.full_name, "error": str(exc)})

    await asyncio.gather(*[_run(s) for s in body.students])
    return result


async def _resend_credentials(body: ResendCredentialsIn) -> dict[str, Any]:
    creds = generate_credentials(body.full_name, body.cedula, settings.institutional_domain)
    login_id, _ = _resolve_login(creds, "student")

    results: dict[str, Any] = {
        "student": body.full_name,
        "cedula": body.cedula,
        "credentials": {**creds, "login_id": login_id},
        "action": "resend",
    }

    try:
        attachments = []
        if body.program_type == "diplomado":
            template_dir = Path(__file__).parent.parent / "static" / "templates"
            pdf_files = [
                template_dir / "2° Acceso a la Plataforma Teams- Instructivo.pdf",
                template_dir / "3° Descargar grabacion en TEAMS - Instructivo.pdf",
            ]
            attachments = [str(f) for f in pdf_files if f.exists()]

        await send_welcome_email(
            to_email=body.personal_email,
            full_name=creds["full_name"],
            institutional_email=creds["email"],
            login_id=login_id,
            password=creds["password"],
            platform=body.platform,
            program_type=body.program_type,
            program_name=body.program_name,
            extra_cc=body.cc or None,
            attachments=attachments if attachments else None,
        )
        results["email"] = "sent"
    except Exception as exc:
        results["email"] = f"error: {exc}"

    return results


@router.post("/resend-credentials", summary="Reenviar credenciales a usuario existente")
async def resend_credentials(body: ResendCredentialsIn) -> dict[str, Any]:
    return await _resend_credentials(body)


@router.post("/bulk-resend", summary="Reenviar credenciales masivas")
async def resend_credentials_bulk(body: BulkResendIn) -> BulkResult:
    result = BulkResult()

    async def _run(student: ResendCredentialsIn):
        try:
            data = await _resend_credentials(student)
            result.succeeded.append(data)
        except Exception as exc:
            result.failed.append({"student": student.full_name, "error": str(exc)})

    await asyncio.gather(*[_run(s) for s in body.students])
    return result


# ── Verificación de cuentas existentes ──────────────────────────────────────

@router.post("/check-account", summary="Verificar si un usuario existe en Canvas y/o Teams")
async def check_account(body: AccountCheckIn) -> dict[str, Any]:
    return await _check_account(body)


@router.post("/bulk-check", summary="Verificar cuentas masivamente")
async def check_accounts_bulk(body: BulkAccountCheckIn) -> dict[str, Any]:
    results = await asyncio.gather(
        *[_check_account(s) for s in body.students],
        return_exceptions=True,
    )
    output = []
    for body_item, res in zip(body.students, results):
        if isinstance(res, Exception):
            output.append({"cedula": body_item.cedula, "full_name": body_item.full_name, "error": str(res)})
        else:
            output.append(res)
    return {
        "total": len(output),
        "found_canvas": sum(1 for r in output if isinstance(r, dict) and r.get("canvas", {}).get("exists")),
        "found_teams": sum(1 for r in output if isinstance(r, dict) and r.get("teams", {}).get("exists")),
        "results": output,
    }


# ── Carga de archivo Excel ──────────────────────────────────────────────────────

@router.get("/template/crear", summary="Descargar plantilla para crear usuarios")
async def template_crear():
    """Descarga plantilla Excel para crear usuarios masivamente."""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["Nombre Completo", "Cedula", "Email Personal", "Rol", "Plataforma", "Programa", "Nombre del Programa"]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="5A67D8", end_color="5A67D8", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    ws.append(["Karen Gonzalez", "6868066", "karen@gmail.com", "student", "both", "grado", ""])
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_crear_usuarios.xlsx"}
    )


@router.post("/bulk-file", summary="Crear usuarios desde archivo Excel")
async def bulk_file_create(file: UploadFile = File(...)) -> BulkResult:
    """Lee archivo Excel y crea usuarios masivamente."""
    import openpyxl
    import io

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")

    result = BulkResult()
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos (mínimo 1 fila después del encabezado)")

    async def _create_from_row(row):
        try:
            if not row[0] or not row[1]:
                return {"error": "Nombre y cédula requeridos"}

            student_data = CreateStudentIn(
                full_name=str(row[0]).strip(),
                cedula=str(row[1]).strip(),
                personal_email=str(row[2]).strip() if row[2] else "",
                role=str(row[3]).strip().lower() if row[3] else "student",
                platform=str(row[4]).strip().lower() if row[4] else "both",
                program_type=str(row[5]).strip().lower() if row[5] else "grado",
                program_name=str(row[6]).strip() if row[6] else "",
                send_email=True,
                cc=[],
            )
            data = await _create_student(student_data)
            result.succeeded.append(data)
        except Exception as exc:
            result.failed.append({
                "student": str(row[0]) if row and row[0] else "?",
                "error": str(exc)[:200],
            })

    await asyncio.gather(*[_create_from_row(row) for row in rows])
    return result
