"""Canvas attendance report processing endpoints.

IMPORTANTE: Los reportes de asistencia NO usan caché para garantizar que
siempre se muestren datos frescos. Se actualizan cada vez que se accede,
permitiendo cambios mensuales en las asistencias registradas.
"""
import logging
from pathlib import Path
from datetime import datetime
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import tempfile
from collections import defaultdict

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/canvas/attendance", tags=["Canvas · Attendance"])

# NOTE: Attendance reports intentionally do NOT use caching to ensure fresh data
# is always displayed, allowing for monthly attendance updates


def parse_attendance_csv(file_path: str | Path) -> dict:
    """Parse Canvas attendance CSV report.

    Expected structure:
    Course ID,SIS Course ID,Course Code,Course Name,Section Name,Section ID,SIS Section ID,
    Teacher ID,Teacher Name,Student ID,Student Name,Class Date,Attendance,Timestamp

    Each row is a student attendance record for a specific class date.
    """
    try:
        file_path = Path(file_path)
        logger.info(f"Parseando archivo CSV de asistencia: {file_path}")

        # Maps to collect data
        course_info = {}
        students_set = {}  # {student_id: student_name}
        dates_set = set()  # unique class dates
        attendance_matrix = defaultdict(dict)  # {student_id: {date: status}}

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):  # start=2 because row 1 is header
                try:
                    # Extract data
                    course_id = row.get('Course ID')
                    course_name = row.get('Course Name', '')
                    teacher_name = row.get('Teacher Name', '')
                    student_id = row.get('Student ID', '').strip()
                    student_name = row.get('Student Name', '').strip()
                    class_date = row.get('Class Date', '').strip()
                    attendance_status = row.get('Attendance', '').strip().lower()

                    # Store course info (use first row's data)
                    if not course_info:
                        course_info = {
                            'id': course_id,
                            'name': course_name,
                            'instructor': teacher_name
                        }

                    # Skip rows without required data
                    if not student_id or not class_date:
                        continue

                    # Store student info
                    if student_id not in students_set:
                        students_set[student_id] = student_name

                    # Collect unique dates
                    dates_set.add(class_date)

                    # Normalize attendance status
                    status_map = {
                        'present': 'P',
                        'absent': 'A',
                        'late': 'L',
                        'excused': 'E',
                        'unmarked': '',
                    }
                    normalized_status = status_map.get(attendance_status, attendance_status)

                    # Store attendance
                    attendance_matrix[student_id][class_date] = normalized_status

                except Exception as e:
                    logger.warning(f"Error en fila {row_num}: {e}")
                    continue

        # Sort students and dates
        students = [
            {"id": int(sid) if sid.isdigit() else sid, "name": students_set[sid]}
            for sid in sorted(students_set.keys(), key=lambda x: int(x) if x.isdigit() else 0)
        ]
        sorted_dates = sorted(dates_set)

        logger.info(f"CSV: {len(students)} estudiantes, {len(sorted_dates)} fechas")
        logger.info(f"Curso: {course_info.get('name')}, Docente: {course_info.get('instructor')}")

        return {
            "course": {
                "id": int(course_info.get('id', 0)) if course_info.get('id', '').isdigit() else course_info.get('id'),
                "name": course_info.get('name', 'Unknown'),
                "instructor": course_info.get('instructor', 'Unknown'),
            },
            "attendance_dates": sorted_dates,
            "students": students,
            "attendance": {
                str(sid): {date: attendance_matrix[sid].get(date, '') for date in sorted_dates}
                for sid in students_set.keys()
            },
            "summary": {
                "total_students": len(students),
                "total_dates": len(sorted_dates),
                "source": "CSV"
            }
        }

    except Exception as exc:
        logger.error(f"Error parseando archivo CSV: {exc}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(exc)}")


def parse_attendance_excel(file_path: str | Path) -> dict:
    """Parse Canvas attendance Excel report.

    Expected structure:
    - Row 1: Course name
    - Row 2: Instructor name
    - Row 3: Empty
    - Row 4: Dates (Col C onwards)
    - Col A: Student IDs
    - Col B: Student names
    - Col C+: Attendance status ("Presente", "Ausente", etc.)
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        logger.info(f"Parseando archivo de asistencia: {file_path}")

        # Extract metadata
        course_name = ws['A1'].value or "Unknown"
        instructor_line = ws['A2'].value or "Docente: Unknown"
        instructor_name = instructor_line.replace("Docente: ", "").strip() if "Docente:" in instructor_line else instructor_line

        # Extract dates from Row 4 (starting from Col C)
        dates = []
        for col_num in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=4, column=col_num).value
            if cell_val:
                # Parse date - could be string like "23/3/2026" or datetime object
                date_str = str(cell_val).strip()
                dates.append(date_str)

        logger.info(f"Cursos: {course_name}, Docente: {instructor_name}, Fechas: {len(dates)}")

        # Extract students and attendance data
        students = []
        attendance_matrix = {}

        for row_num in range(5, ws.max_row + 1):
            student_id = ws.cell(row=row_num, column=1).value
            student_name = ws.cell(row=row_num, column=2).value

            if not student_id or not student_name:
                continue

            student_id = int(student_id) if isinstance(student_id, (int, float)) else str(student_id)

            students.append({
                "id": student_id,
                "name": str(student_name).strip(),
            })

            # Extract attendance for each date
            attendance_record = {}
            for col_num, date_str in enumerate(dates, start=3):
                cell_val = ws.cell(row=row_num, column=col_num).value
                status = str(cell_val).strip() if cell_val else ""

                # Normalize status to single letter
                if status.lower() in ["presente", "present", "p"]:
                    attendance_record[date_str] = "P"
                elif status.lower() in ["ausente", "absent", "a"]:
                    attendance_record[date_str] = "A"
                elif status.lower() in ["retardo", "late", "tardío", "l"]:
                    attendance_record[date_str] = "L"
                elif status.lower() in ["excusado", "excused", "e"]:
                    attendance_record[date_str] = "E"
                else:
                    attendance_record[date_str] = status

            attendance_matrix[str(student_id)] = attendance_record
            logger.debug(f"Estudiante {student_id} ({student_name}): {attendance_record}")

        return {
            "course": {
                "name": course_name,
                "instructor": instructor_name,
            },
            "dates": dates,
            "students": students,
            "attendance": attendance_matrix,
            "summary": {
                "total_students": len(students),
                "total_dates": len(dates),
                "source": "Excel"
            }
        }

    except Exception as exc:
        logger.error(f"Error parseando archivo Excel: {exc}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(exc)}")


@router.post("/upload", summary="Cargar reporte de asistencia Excel")
async def upload_attendance_report(file: UploadFile = File(...)):
    """Upload and parse Canvas attendance Excel report."""

    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(status_code=400, detail="Archivo debe ser Excel (.xlsx, .xls) o CSV")

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        logger.info(f"Archivo cargado: {file.filename} -> {tmp_path}")

        # Parse the file
        result = parse_attendance_excel(tmp_path)

        # Cleanup
        Path(tmp_path).unlink()

        return result

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error cargando archivo: {exc}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(exc)}")


@router.get("/from-csv/{course_id}", summary="Leer reporte de asistencia (CSV o API)")
async def load_attendance_from_csv(course_id: str):
    """Load attendance report from CSV file if available, otherwise from Canvas API.

    Strategy:
    1. First, try to load from CSV file (Carpeta de plantillas/{course_id}.csv)
    2. If CSV doesn't exist, fetch from Canvas API (/canvas/courses/{course_id}/attendance)
    3. If both fail, return empty report with course info
    """

    try:
        from app.services import canvas_client as canvas

        # Step 1: Try to load from CSV
        template_dir = Path("Carpeta de plantillas")
        csv_file = template_dir / f"{course_id}.csv"

        if template_dir.exists() and csv_file.exists():
            logger.info(f"Cargando CSV de asistencia: {csv_file}")
            result = parse_attendance_csv(csv_file)
            result['source'] = 'CSV'
            return result

        # Step 2: Fallback to Canvas API
        logger.info(f"CSV no encontrado, obteniendo datos de Canvas API para curso {course_id}")

        try:
            # Get course info
            course = await canvas.get(f"/courses/{course_id}")
            course_name = course.get("name", f"Curso {course_id}")

            # Get instructor
            enrollments = await canvas.paginate_limited(
                f"/courses/{course_id}/enrollments",
                {"per_page": 100, "type": "TeacherEnrollment"},
                max_records=100
            )
            instructor_name = ""
            if enrollments:
                teacher = enrollments[0].get("user", {})
                instructor_name = teacher.get("name", "")

            # Get students
            all_enrollments = await canvas.paginate_limited(
                f"/courses/{course_id}/enrollments",
                {"per_page": 100},
                max_records=1000
            )

            student_enrollments = [
                e for e in all_enrollments
                if e.get("type") == "StudentEnrollment" and
                   e.get("enrollment_state") in ["active", "completed"]
            ]

            students = []
            student_ids = set()
            for enrollment in student_enrollments:
                user_id = enrollment.get("user_id")
                user = enrollment.get("user", {})
                if user_id:
                    students.append({
                        "id": user_id,
                        "name": user.get("name", f"Usuario {user_id}"),
                        "login": user.get("login", ""),
                    })
                    student_ids.add(user_id)

            logger.info(f"Obtenidos {len(students)} estudiantes del curso {course_id}")

            return {
                "course": {
                    "id": int(course_id),
                    "name": course_name,
                    "instructor": instructor_name,
                },
                "attendance_dates": [],
                "students": students,
                "attendance": {str(uid): {} for uid in student_ids},
                "summary": {
                    "total_students": len(students),
                    "total_dates": 0,
                    "source": "Canvas API (sin asistencias)",
                    "note": "Los datos de asistencia están disponibles en el CSV. Carga el CSV del curso para obtener detalles de asistencia."
                }
            }

        except Exception as api_exc:
            logger.warning(f"Error obteniendo datos de API para curso {course_id}: {api_exc}")
            # Return minimal response
            return {
                "course": {
                    "id": int(course_id) if course_id.isdigit() else course_id,
                    "name": f"Curso {course_id}",
                    "instructor": "Desconocido",
                },
                "attendance_dates": [],
                "students": [],
                "attendance": {},
                "summary": {
                    "total_students": 0,
                    "total_dates": 0,
                    "source": "Error",
                    "error": "No se pudieron obtener datos de asistencia"
                }
            }

    except Exception as exc:
        logger.error(f"Error cargando asistencia para curso {course_id}: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error procesando asistencia: {str(exc)}")


@router.get("/from-template/{course_id}", summary="Leer reporte de asistencia desde plantillas")
async def load_attendance_from_template(course_id: str):
    """Load attendance report from template folder if it exists.

    Looks for Excel files matching the pattern in:
    Carpeta de plantillas/REPORTE_ASISTENCIA_*.xlsx
    """

    try:
        template_dir = Path("Carpeta de plantillas")

        if not template_dir.exists():
            raise HTTPException(status_code=404, detail="Carpeta de plantillas no encontrada")

        # Search for attendance report files
        attendance_files = list(template_dir.glob("REPORTE_ASISTENCIA_*.xlsx"))

        if not attendance_files:
            raise HTTPException(status_code=404, detail=f"No se encontraron reportes de asistencia en {template_dir}")

        logger.info(f"Encontrados {len(attendance_files)} reportes de asistencia")

        # Use the most recent file
        latest_file = max(attendance_files, key=lambda p: p.stat().st_mtime)
        logger.info(f"Usando archivo: {latest_file}")

        # Parse the file
        result = parse_attendance_excel(latest_file)

        return result

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error cargando plantilla: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error procesando plantilla: {str(exc)}")


@router.get("/templates", summary="Listar reportes disponibles en plantillas")
async def list_attendance_templates():
    """List all attendance report files in template folder."""

    try:
        template_dir = Path("Carpeta de plantillas")

        if not template_dir.exists():
            return {"templates": [], "total": 0}

        # Find all attendance report files
        files = list(template_dir.glob("REPORTE_ASISTENCIA_*.xlsx"))

        file_info = []
        for f in files:
            stat = f.stat()
            file_info.append({
                "filename": f.name,
                "size": stat.st_size,
                "modified": stat.st_mtime,
                "path": str(f)
            })

        # Sort by modified date, newest first
        file_info.sort(key=lambda x: x["modified"], reverse=True)

        logger.info(f"Encontrados {len(file_info)} reportes de asistencia")

        return {
            "templates": file_info,
            "total": len(file_info),
            "directory": str(template_dir)
        }

    except Exception as exc:
        logger.error(f"Error listando plantillas: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error listando plantillas: {str(exc)}")


def generate_attendance_excel(report: dict) -> Path:
    """Generate Excel file from attendance report data.

    Format matches Canvas attendance reports:
    - Row 1: Course name
    - Row 2: Instructor name
    - Row 3: Empty
    - Row 4: Dates (starting from Col C)
    - Col A: Student IDs
    - Col B: Student names
    - Col C+: Attendance status (P/A/L/E)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report.get('courseId', 'Attendance')}"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    date_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    date_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Course name
    ws['A1'] = report['course']['name']
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')

    # Row 2: Instructor
    ws['A2'] = f"Docente: {report['course']['instructor']}"
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:L2')

    # Row 3: Empty
    ws['A3'] = ""

    # Row 4: Dates (starting from Col C)
    dates = report['attendance_dates']
    for col_idx, date in enumerate(dates, start=3):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = date
        cell.fill = date_fill
        cell.font = date_font
        cell.alignment = center_alignment
        cell.border = border

    # Rows 5+: Students and attendance
    students = report['students']
    attendance = report['attendance']

    for row_idx, student in enumerate(students, start=5):
        # Column A: Student ID
        id_cell = ws.cell(row=row_idx, column=1)
        id_cell.value = student['id']
        id_cell.border = border

        # Column B: Student name
        name_cell = ws.cell(row=row_idx, column=2)
        name_cell.value = student['name']
        name_cell.border = border

        # Columns C+: Attendance
        student_id = str(student['id'])
        student_attendance = attendance.get(student_id, {})

        for col_idx, date in enumerate(dates, start=3):
            cell = ws.cell(row=row_idx, column=col_idx)
            status = student_attendance.get(date, '')

            # Map status codes to labels
            status_map = {
                'P': 'Presente',
                'A': 'Ausente',
                'L': 'Retardo',
                'E': 'Excusado',
                '': ''
            }
            cell.value = status_map.get(status, status)
            cell.alignment = center_alignment
            cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    for col_idx in range(3, 3 + len(dates)):
        ws.column_dimensions[chr(64 + col_idx)].width = 14

    # Set row height
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 16

    # Save to temporary file
    temp_file = Path(tempfile.gettempdir()) / f"asistencia_{report['courseId']}.xlsx"
    wb.save(temp_file)
    logger.info(f"Excel generado: {temp_file}")

    return temp_file


@router.get("/download-excel/{course_id}", summary="Descargar reporte como Excel")
async def download_attendance_excel(course_id: str):
    """Download attendance report as Excel file.

    First tries to use CSV data if available, then falls back to API data.
    """

    try:
        # Get report data (from CSV or API)
        # We call the load_attendance_from_csv endpoint which handles both sources
        from app.routers.canvas.attendance import load_attendance_from_csv

        report = await load_attendance_from_csv(course_id)
        report['courseId'] = course_id

        # Generate Excel
        excel_file = generate_attendance_excel(report)

        # Return file
        filename = f"Asistencia_{report['course']['name'].replace('/', '_')[:40]}.xlsx"

        return FileResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error descargando Excel para curso {course_id}: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(exc)}")
